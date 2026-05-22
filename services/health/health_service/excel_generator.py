import io

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side, Alignment
from openpyxl.utils import get_column_letter


FONT_NAME = "Microsoft YaHei"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name=FONT_NAME, size=10)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, num_cols: int) -> None:
    """Apply professional header styling and frozen top row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_cell_style(ws, num_rows: int, num_cols: int) -> None:
    """Apply border and font to all data cells."""
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, num_cols: int) -> None:
    """Set reasonable column widths based on header and content."""
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    # Rough width: CJK chars count as ~2.2, ASCII as 1
                    val = str(cell_val)
                    width = sum(2.2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, width)
        ws.column_dimensions[letter].width = min(max_len + 4, 40)


def _save_workbook(wb: Workbook) -> bytes:
    """Save workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# FMEA Excel
# ---------------------------------------------------------------------------

FMEA_HEADERS = [
    "ID",
    "设备类型",
    "部件",
    "故障模式",
    "故障影响",
    "严重度(S)",
    "发生度(O)",
    "检测度(D)",
    "RPN",
    "缓解措施",
    "症状特征",
]


def generate_fmea_excel(items: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FMEA知识库"
    ws.freeze_panes = "A2"

    ws.append(FMEA_HEADERS)
    for item in items:
        symptoms = item.get("symptoms")
        symptoms_str = str(symptoms) if symptoms else ""
        ws.append([
            item.get("id"),
            item.get("equipment_type"),
            item.get("component"),
            item.get("failure_mode"),
            item.get("effects"),
            item.get("severity"),
            item.get("occurrence"),
            item.get("detection"),
            item.get("rpn"),
            item.get("mitigation"),
            symptoms_str,
        ])

    num_rows = ws.max_row
    num_cols = len(FMEA_HEADERS)
    _apply_header_style(ws, num_cols)
    _apply_cell_style(ws, num_rows, num_cols)
    _auto_width(ws, num_cols)

    return _save_workbook(wb)


# ---------------------------------------------------------------------------
# RUL Excel
# ---------------------------------------------------------------------------

RUL_HEADERS = [
    "设备ID",
    "部件",
    "剩余寿命(小时)",
    "置信下限(小时)",
    "置信上限(小时)",
    "退化模型",
]


def generate_rul_excel(items: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "剩余寿命预测"
    ws.freeze_panes = "A2"

    ws.append(RUL_HEADERS)
    for item in items:
        ws.append([
            item.get("equipment_id"),
            item.get("component"),
            item.get("predicted_hours"),
            item.get("ci_lo"),
            item.get("ci_hi"),
            item.get("degradation_model"),
        ])

    num_rows = ws.max_row
    num_cols = len(RUL_HEADERS)
    _apply_header_style(ws, num_cols)
    _apply_cell_style(ws, num_rows, num_cols)
    _auto_width(ws, num_cols)

    return _save_workbook(wb)


# ---------------------------------------------------------------------------
# Diagnosis Excel
# ---------------------------------------------------------------------------

DIAGNOSIS_HEADERS = [
    "ID",
    "设备ID",
    "根本原因",
    "置信度",
    "严重度(1-5)",
    "认证等级(1-4)",
    "时间戳",
]


def generate_diagnosis_excel(items: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "故障诊断记录"
    ws.freeze_panes = "A2"

    ws.append(DIAGNOSIS_HEADERS)
    for item in items:
        ws.append([
            item.get("id"),
            item.get("equipment_id"),
            item.get("root_cause"),
            item.get("confidence"),
            item.get("severity"),
            item.get("cert_level"),
            item.get("timestamp"),
        ])

    num_rows = ws.max_row
    num_cols = len(DIAGNOSIS_HEADERS)
    _apply_header_style(ws, num_cols)
    _apply_cell_style(ws, num_rows, num_cols)
    _auto_width(ws, num_cols)

    return _save_workbook(wb)


# ---------------------------------------------------------------------------
# Health Report Excel
# ---------------------------------------------------------------------------


def generate_health_report_excel(equipment_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"设备{equipment_data.get('equipment_id', '')}健康报告"
    ws.freeze_panes = "A2"

    # Build a structured report
    ws.append(["项目", "值"])
    equipment_id = equipment_data.get("equipment_id", "")
    ws.append(["设备ID", equipment_id])
    ws.append(["综合健康评分", equipment_data.get("overall_score")])

    trend_info = equipment_data.get("trend", {})
    ws.append(["趋势方向", trend_info.get("direction")])
    ws.append(["趋势斜率", trend_info.get("slope")])

    ws.append(["振动区", equipment_data.get("vibration_zone")])
    ws.append([])  # blank row

    # Component scores
    ws.append(["部件评分", ""])
    component_scores = equipment_data.get("component_scores", {})
    for comp, score in component_scores.items():
        ws.append([comp, score])

    ws.append([])

    # RUL section
    ruling = equipment_data.get("latest_rul")
    if ruling:
        ws.append(["最新剩余寿命预测", ""])
        ws.append(["部件", ruling.get("component")])
        ws.append(["预测剩余小时", ruling.get("predicted_hours")])
        ws.append(["置信下限", ruling.get("ci_lo")])
        ws.append(["置信上限", ruling.get("ci_hi")])

    ws.append([])

    # Degradation history
    deg_history = equipment_data.get("degradation_history", [])
    if deg_history:
        ws.append(["退化历史", ""])
        ws.append(["日期", "评分"])

        # Style the sub-header row
        sub_header_row = ws.max_row
        for col in range(1, 3):
            cell = ws.cell(row=sub_header_row, column=col)
            cell.font = Font(name=FONT_NAME, bold=True, size=10)
            cell.border = THIN_BORDER

        for entry in deg_history:
            ws.append([entry.get("date"), entry.get("score")])

    ws.append([])

    # Recent diagnoses
    diagnoses = equipment_data.get("recent_diagnoses", [])
    if diagnoses:
        ws.append(["近期诊断", ""])
        ws.append(["诊断ID", "根本原因", "置信度", "日期"])
        sub_header_row2 = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=sub_header_row2, column=col)
            cell.font = Font(name=FONT_NAME, bold=True, size=10)
            cell.border = THIN_BORDER

        for d in diagnoses:
            ws.append([d.get("id"), d.get("root_cause"), d.get("confidence"), d.get("date")])

    # Apply professional header styling only to the very first header row
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    num_rows = ws.max_row
    # Apply cell borders to all populated rows (skip section headers which already have borders)
    for row in range(2, num_rows + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

    _auto_width(ws, 4)

    return _save_workbook(wb)

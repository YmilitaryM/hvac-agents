import io
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Query
from fastapi.responses import Response
from openpyxl import load_workbook

from ..excel_generator import (
    generate_fmea_excel,
    generate_rul_excel,
    generate_diagnosis_excel,
    generate_health_report_excel,
)

router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ---------------------------------------------------------------------------
# Demo data — mirrors existing endpoint patterns
# ---------------------------------------------------------------------------

FMEA_DEMO = [
    {
        "id": 1,
        "equipment_type": "离心式冷水机组",
        "component": "压缩机",
        "failure_mode": "轴承磨损",
        "effects": "振动加剧，效率下降，可能导致压缩机卡死",
        "severity": 7,
        "occurrence": 4,
        "detection": 3,
        "rpn": 84,
        "mitigation": "定期振动监测，每3个月更换润滑油",
        "symptoms": {"vibration_rms": ">7.0", "temp_rise": ">10"},
    },
    {
        "id": 2,
        "equipment_type": "离心式冷水机组",
        "component": "压缩机",
        "failure_mode": "叶轮腐蚀",
        "effects": "流量下降，能耗上升",
        "severity": 6,
        "occurrence": 3,
        "detection": 4,
        "rpn": 72,
        "mitigation": "水质处理，定期检查叶轮表面",
        "symptoms": {"flow_drop": ">15%", "power_increase": ">10%"},
    },
    {
        "id": 3,
        "equipment_type": "离心式冷水机组",
        "component": "电机",
        "failure_mode": "绝缘老化",
        "effects": "电机过热，可能烧毁",
        "severity": 9,
        "occurrence": 2,
        "detection": 4,
        "rpn": 72,
        "mitigation": "定期绝缘电阻测试，保持通风良好",
        "symptoms": {"insulation_resistance": "<1MΩ", "temp_rise": ">15"},
    },
    {
        "id": 4,
        "equipment_type": "冷却塔",
        "component": "风扇",
        "failure_mode": "轴承卡滞",
        "effects": "冷却能力下降，可能停机",
        "severity": 5,
        "occurrence": 5,
        "detection": 2,
        "rpn": 50,
        "mitigation": "定期润滑，振动监测",
        "symptoms": {"vibration_increase": ">50%", "noise": "异常噪音"},
    },
    {
        "id": 5,
        "equipment_type": "水泵",
        "component": "机械密封",
        "failure_mode": "密封泄漏",
        "effects": "水损增加，可能损坏轴承",
        "severity": 4,
        "occurrence": 6,
        "detection": 2,
        "rpn": 48,
        "mitigation": "定期检查密封面，及时更换",
        "symptoms": {"leak_rate": ">10ml/min", "pressure_drop": ">0.5bar"},
    },
]

RUL_DEMO = [
    {
        "equipment_id": 1,
        "component": "轴承",
        "predicted_hours": 2000,
        "ci_lo": 1500,
        "ci_hi": 2500,
        "degradation_model": "weibull",
    },
    {
        "equipment_id": 1,
        "component": "叶轮",
        "predicted_hours": 8000,
        "ci_lo": 6500,
        "ci_hi": 9500,
        "degradation_model": "linear",
    },
    {
        "equipment_id": 2,
        "component": "压缩机",
        "predicted_hours": 5000,
        "ci_lo": 4200,
        "ci_hi": 5800,
        "degradation_model": "exp",
    },
    {
        "equipment_id": 2,
        "component": "电机绕组",
        "predicted_hours": 12000,
        "ci_lo": 10000,
        "ci_hi": 14000,
        "degradation_model": "weibull",
    },
    {
        "equipment_id": 3,
        "component": "风扇轴承",
        "predicted_hours": 3500,
        "ci_lo": 2800,
        "ci_hi": 4200,
        "degradation_model": "exp",
    },
]

DIAGNOSIS_DEMO = [
    {
        "id": 1,
        "equipment_id": 1,
        "root_cause": "轴承磨损",
        "confidence": 0.85,
        "severity": 3,
        "cert_level": 2,
        "timestamp": "2026-05-20T10:30:00",
    },
    {
        "id": 2,
        "equipment_id": 1,
        "root_cause": "不对中",
        "confidence": 0.62,
        "severity": 2,
        "cert_level": 1,
        "timestamp": "2026-05-21T14:00:00",
    },
    {
        "id": 3,
        "equipment_id": 2,
        "root_cause": "压缩机润滑油劣化",
        "confidence": 0.91,
        "severity": 4,
        "cert_level": 3,
        "timestamp": "2026-05-19T08:15:00",
    },
    {
        "id": 4,
        "equipment_id": 3,
        "root_cause": "风扇不平衡",
        "confidence": 0.78,
        "severity": 2,
        "cert_level": 2,
        "timestamp": "2026-05-18T16:45:00",
    },
    {
        "id": 5,
        "equipment_id": 4,
        "root_cause": "机械密封泄漏",
        "confidence": 0.88,
        "severity": 3,
        "cert_level": 2,
        "timestamp": "2026-05-22T09:00:00",
    },
]


# ---------------------------------------------------------------------------
# Download endpoints
# ---------------------------------------------------------------------------

@router.get("/fmea/download")
async def download_fmea_excel():
    data = generate_fmea_excel(FMEA_DEMO)
    return Response(
        content=data,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=fmea_knowledge_base.xlsx"},
    )


@router.get("/rul/download")
async def download_rul_excel():
    data = generate_rul_excel(RUL_DEMO)
    return Response(
        content=data,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=rul_predictions.xlsx"},
    )


@router.get("/diagnosis/download")
async def download_diagnosis_excel():
    data = generate_diagnosis_excel(DIAGNOSIS_DEMO)
    return Response(
        content=data,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=diagnosis_history.xlsx"},
    )


@router.get("/equipment/{equipment_id}/report/download")
async def download_equipment_report(equipment_id: int):
    # Use the same demo pattern as equipment_detail endpoint
    report_data = {
        "equipment_id": equipment_id,
        "overall_score": 85,
        "component_scores": {
            "压缩机": 90,
            "轴承": 78,
            "电机绕组": 88,
            "换热器": 82,
        },
        "trend": {"direction": "稳定", "slope": -0.05},
        "degradation_history": [
            {"date": "2026-05-15", "score": 87},
            {"date": "2026-05-18", "score": 86},
            {"date": "2026-05-21", "score": 85},
        ],
        "latest_rul": {
            "component": "轴承",
            "predicted_hours": 2000,
            "ci_lo": 1500,
            "ci_hi": 2500,
        },
        "recent_diagnoses": [
            {"id": 1, "root_cause": "轻微不对中", "confidence": 0.72, "date": "2026-05-20"},
        ],
        "vibration_zone": "B",
    }
    data = generate_health_report_excel(report_data)
    return Response(
        content=data,
        media_type=XLSX_MIME,
        headers={
            "Content-Disposition": f"attachment; filename=equipment_{equipment_id}_health_report.xlsx"
        },
    )


# ---------------------------------------------------------------------------
# Import endpoint
# ---------------------------------------------------------------------------

# Expected FMEA import columns (Chinese headers)
FMEA_IMPORT_MAPPING = {
    "设备类型": "equipment_type",
    "部件": "component",
    "故障模式": "failure_mode",
    "故障影响": "effects",
    "影响": "effects",  # alternative header
    "严重度(S)": "severity",
    "严重度": "severity",  # alternative header
    "发生度(O)": "occurrence",
    "发生度": "occurrence",  # alternative header
    "检测度(D)": "detection",
    "检测度": "detection",  # alternative header
    "RPN": "rpn",
    "缓解措施": "mitigation",
    "症状特征": "symptoms",
}

REQUIRED_IMPORT_COLUMNS = {"equipment_type", "component", "failure_mode", "severity", "occurrence", "detection"}


@router.post("/fmea/import")
async def import_fmea_excel(file: UploadFile = File(...)):
    # Validate file extension
    if not file.filename or not file.filename.endswith((".xlsx", ".xlsm")):
        return Response(
            content='{"detail": "Invalid file type. Only .xlsx files are accepted."}',
            media_type="application/json",
            status_code=422,
        )

    content = await file.read()

    # Validate ZIP magic bytes
    if content[:4] != b"PK\x03\x04":
        return Response(
            content='{"detail": "File is not a valid Excel file."}',
            media_type="application/json",
            status_code=422,
        )

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return Response(
            content='{"detail": "Cannot parse Excel file."}',
            media_type="application/json",
            status_code=422,
        )

    ws = wb.active
    if ws is None:
        return Response(
            content='{"detail": "Excel file has no active worksheet."}',
            media_type="application/json",
            status_code=422,
        )

    # Read headers from first row
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        return Response(
            content='{"detail": "Excel file is empty."}',
            media_type="application/json",
            status_code=422,
        )

    # Build header mapping
    header_map = {}
    for idx, header in enumerate(headers_row):
        if header and str(header).strip() in FMEA_IMPORT_MAPPING:
            header_map[idx] = FMEA_IMPORT_MAPPING[str(header).strip()]

    # Check required columns
    mapped_fields = set(header_map.values())
    if not REQUIRED_IMPORT_COLUMNS.issubset(mapped_fields):
        missing = REQUIRED_IMPORT_COLUMNS - mapped_fields
        return Response(
            content=f'{{"detail": "Missing required columns: {", ".join(sorted(missing))}"}}',
            media_type="application/json",
            status_code=422,
        )

    # Parse data rows
    records = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue  # skip empty rows

        record = {}
        for idx, value in enumerate(row):
            if idx in header_map and value is not None:
                field = header_map[idx]
                record[field] = value

        # Only include rows that have minimally the required fields
        if REQUIRED_IMPORT_COLUMNS.issubset(record.keys()):
            records.append(record)

    wb.close()

    return {
        "imported_count": len(records),
        "records": records,
    }

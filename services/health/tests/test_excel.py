import io
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from health_service.main import app
from health_service.excel_generator import (
    generate_fmea_excel,
    generate_rul_excel,
    generate_diagnosis_excel,
    generate_health_report_excel,
)


# ---------------------------------------------------------------------------
# Unit tests for Excel generators
# ---------------------------------------------------------------------------

FMEA_SAMPLE = [
    {
        "id": 1,
        "equipment_type": "离心式冷水机组",
        "component": "压缩机",
        "failure_mode": "轴承磨损",
        "effects": "振动加剧，效率下降",
        "severity": 7,
        "occurrence": 4,
        "detection": 3,
        "rpn": 84,
        "mitigation": "定期振动监测，每3个月更换润滑油",
        "symptoms": {"vibration_rms": ">7.0", "temp_rise": ">10"},
    },
    {
        "id": 2,
        "equipment_type": "离心式冷水机组",
        "component": "电机",
        "failure_mode": "绝缘老化",
        "effects": "电机过热，可能烧毁",
        "severity": 9,
        "occurrence": 2,
        "detection": 4,
        "rpn": 72,
        "mitigation": "定期绝缘电阻测试",
        "symptoms": {"insulation_resistance": "<1MΩ", "temp_rise": ">15"},
    },
]

RUL_SAMPLE = [
    {
        "equipment_id": 1,
        "component": "轴承",
        "predicted_hours": 2000,
        "ci_lo": 1500,
        "ci_hi": 2500,
        "degradation_model": "weibull",
    },
    {
        "equipment_id": 2,
        "component": "压缩机",
        "predicted_hours": 5000,
        "ci_lo": 4200,
        "ci_hi": 5800,
        "degradation_model": "exp",
    },
]

DIAGNOSIS_SAMPLE = [
    {
        "id": 1,
        "equipment_id": 1,
        "root_cause": "轴承磨损",
        "confidence": 0.85,
        "severity": 3,
        "cert_level": 2,
        "timestamp": "2026-05-20T10:30:00",
    },
    {
        "id": 2,
        "equipment_id": 1,
        "root_cause": "不对中",
        "confidence": 0.62,
        "severity": 2,
        "cert_level": 1,
        "timestamp": "2026-05-21T14:00:00",
    },
]

EQUIPMENT_REPORT_SAMPLE = {
    "equipment_id": 1,
    "overall_score": 85,
    "component_scores": {
        "压缩机": 90,
        "轴承": 78,
        "电机绕组": 88,
        "换热器": 82,
    },
    "trend": {"direction": "稳定", "slope": -0.05},
    "degradation_history": [
        {"date": "2026-05-15", "score": 87},
        {"date": "2026-05-18", "score": 86},
        {"date": "2026-05-21", "score": 85},
    ],
    "latest_rul": {
        "component": "轴承",
        "predicted_hours": 2000,
        "ci_lo": 1500,
        "ci_hi": 2500,
    },
    "recent_diagnoses": [
        {"id": 1, "root_cause": "轻微不对中", "confidence": 0.72, "date": "2026-05-20"},
    ],
    "vibration_zone": "B",
}


class TestFMEAExcel:
    def test_generates_valid_xlsx_bytes(self):
        data = generate_fmea_excel(FMEA_SAMPLE)
        assert isinstance(data, bytes)
        # XLSX files start with ZIP magic bytes (PK\x03\x04)
        assert data[:4] == b"PK\x03\x04"

    def test_can_be_reopened_with_openpyxl(self):
        data = generate_fmea_excel(FMEA_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws is not None

    def test_chinese_headers_present(self):
        data = generate_fmea_excel(FMEA_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "故障模式" in headers
        assert "部件" in headers
        assert any("严重度" in (h or "") for h in headers)

    def test_data_rows_written(self):
        data = generate_fmea_excel(FMEA_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=3, column=1).value is not None


class TestRULExcel:
    def test_generates_valid_xlsx_bytes(self):
        data = generate_rul_excel(RUL_SAMPLE)
        assert data[:4] == b"PK\x03\x04"

    def test_can_be_reopened_with_openpyxl(self):
        data = generate_rul_excel(RUL_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active is not None

    def test_chinese_headers_present(self):
        data = generate_rul_excel(RUL_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "部件" in headers
        assert any("剩余寿命" in (h or "") for h in headers)

    def test_data_rows_written(self):
        data = generate_rul_excel(RUL_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value is not None


class TestDiagnosisExcel:
    def test_generates_valid_xlsx_bytes(self):
        data = generate_diagnosis_excel(DIAGNOSIS_SAMPLE)
        assert data[:4] == b"PK\x03\x04"

    def test_can_be_reopened_with_openpyxl(self):
        data = generate_diagnosis_excel(DIAGNOSIS_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active is not None

    def test_chinese_headers_present(self):
        data = generate_diagnosis_excel(DIAGNOSIS_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "根本原因" in headers
        assert "置信度" in headers

    def test_data_rows_written(self):
        data = generate_diagnosis_excel(DIAGNOSIS_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value is not None


class TestHealthReportExcel:
    def test_generates_valid_xlsx_bytes(self):
        data = generate_health_report_excel(EQUIPMENT_REPORT_SAMPLE)
        assert data[:4] == b"PK\x03\x04"

    def test_can_be_reopened_with_openpyxl(self):
        data = generate_health_report_excel(EQUIPMENT_REPORT_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active is not None

    def test_contains_equipment_id(self):
        data = generate_health_report_excel(EQUIPMENT_REPORT_SAMPLE)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Search all cells for the equipment_id
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell == 1 or str(cell) == "1":
                    found = True
                    break
        assert found, "Equipment ID should appear in the report"


# ---------------------------------------------------------------------------
# Integration tests for download / import endpoints
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def client():
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as c:
        yield c


@pytest.mark.asyncio
async def test_fmea_download_endpoint(client):
    r = await client.get("/api/health/fmea/download")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert r.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_rul_download_endpoint(client):
    r = await client.get("/api/health/rul/download")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert r.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_diagnosis_download_endpoint(client):
    r = await client.get("/api/health/diagnosis/download")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert r.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_equipment_report_download_endpoint(client):
    r = await client.get("/api/health/equipment/1/report/download")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert r.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_fmea_import_valid_excel(client):
    # Build a minimal FMEA Excel file for import
    import io as _io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "FMEA"
    ws.append(["设备类型", "部件", "故障模式", "影响", "严重度", "发生度", "检测度", "RPN", "缓解措施"])
    ws.append(["离心式冷水机组", "压缩机", "轴承磨损", "振动加剧", 7, 4, 3, 84, "定期监测"])
    ws.append(["离心式冷水机组", "电机", "绝缘老化", "过热", 9, 2, 4, 72, "绝缘测试"])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {"file": ("fmea.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await client.post("/api/health/fmea/import", files=files)
    assert r.status_code == 200
    data = r.json()
    assert data["imported_count"] == 2
    assert len(data["records"]) == 2
    assert data["records"][0]["component"] == "压缩机"


@pytest.mark.asyncio
async def test_fmea_import_rejects_invalid_file(client):
    files = {"file": ("not_excel.txt", io.BytesIO(b"hello world"), "text/plain")}
    r = await client.post("/api/health/fmea/import", files=files)
    assert r.status_code == 422

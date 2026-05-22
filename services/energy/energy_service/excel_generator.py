"""Excel export generators for the energy service.

Produces .xlsx bytes using openpyxl with Chinese column headers,
professional styling, and auto-width columns.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Shared styling constants ─────────────────────────────────────────────

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")


def _style_header(ws, num_cols: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_cells(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply border and alignment to data cells."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Rough CJK width estimation: CJK chars count as ~2
                val = str(cell.value)
                length = sum(2 if "一" <= ch <= "鿿" or "　" <= ch <= "〿" else 1 for ch in val)
                if length > max_len:
                    max_len = length
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


def _to_bytes(wb) -> bytes:
    """Serialize a workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Energy Report Excel ──────────────────────────────────────────────────

REPORT_HEADERS = ["编号", "上报周期", "报告类型", "创建时间"]


def generate_energy_report_excel(period: str, data: dict[str, Any]) -> bytes:
    """Generate an energy report list Excel file.

    Args:
        period: Report period (day / month / year).
        data: Dict with an "items" key containing a list of report dicts.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "能源报告"

    # Title row
    period_labels = {"day": "日报", "month": "月报", "year": "年报"}
    period_label = period_labels.get(period, period)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"能源报告列表 — {period_label}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    header_row = 2
    for col_idx, header in enumerate(REPORT_HEADERS, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header(ws, len(REPORT_HEADERS))

    # Data rows
    items: list[dict] = data.get("items", [])
    for row_idx, item in enumerate(items, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=item.get("id"))
        ws.cell(row=row_idx, column=2, value=item.get("period"))
        ws.cell(row=row_idx, column=3, value=item.get("report_type"))
        ws.cell(row=row_idx, column=4, value=item.get("created_at"))

    if items:
        _style_data_cells(ws, header_row + 1, header_row + len(items), len(REPORT_HEADERS))

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return _to_bytes(wb)


# ── Measurement & Verification Excel ─────────────────────────────────────

MV_FIELDS = [
    ("厂站编号", "plant_id"),
    ("基准能耗 (kWh)", "baseline_energy_kwh"),
    ("实际能耗 (kWh)", "actual_energy_kwh"),
    ("节能量 (kWh)", "savings_kwh"),
    ("节能率 (%)", "savings_pct"),
    ("不确定度 (%)", "uncertainty_pct"),
    ("CV(RMSE) (%)", "cv_rmse_pct"),
    ("NMBE (%)", "nmbe_pct"),
    ("ASHRAE G14 合规", "compliant_ashrae_g14"),
    ("GB 28750 合规", "compliant_gb28750"),
    ("等效标煤 (吨)", "coal_equivalent_tons"),
    ("碳减排量 (kg)", "carbon_reduction_kg"),
]


def generate_mv_excel(data: dict[str, Any]) -> bytes:
    """Generate a Measurement & Verification Excel file.

    Args:
        data: Dict of MV verification results (matching the /mv/verify endpoint response).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M&V 节能量验证"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws.cell(row=1, column=1, value="节能量测量与验证 (M&V) 报告")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    ws.cell(row=2, column=1, value="指标")
    ws.cell(row=2, column=2, value="数值")
    _style_header(ws, 2)

    # Data rows
    for row_idx, (label, key) in enumerate(MV_FIELDS, start=3):
        ws.cell(row=row_idx, column=1, value=label)
        value = data.get(key)
        if isinstance(value, bool):
            value = "是" if value else "否"
        ws.cell(row=row_idx, column=2, value=value)

    _style_data_cells(ws, 3, 2 + len(MV_FIELDS), 2)

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=3, column=1)

    return _to_bytes(wb)


# ── Comparison Excel ─────────────────────────────────────────────────────

COMPARISON_HEADERS = [
    "指标",
    "当前值",
    "上期值",
    "环比变化 (%)",
    "同比变化 (%)",
]


def generate_comparison_excel(data: dict[str, Any]) -> bytes:
    """Generate a YoY/MoM comparison Excel file.

    Args:
        data: Dict with current, previous, mom_change_pct, yoy_change_pct keys.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "能耗对比"

    # Title
    period_label = {"day": "日报", "month": "月报", "year": "年报"}.get(data.get("period", "month"), "月报")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COMPARISON_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"能耗对比报告 — {period_label}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    header_row = 2
    for col_idx, header in enumerate(COMPARISON_HEADERS, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header(ws, len(COMPARISON_HEADERS))

    # Data rows — one row per metric
    metric_labels = {
        "total_kwh": "总能耗 (kWh)",
        "avg_cop": "平均COP",
        "avg_power_kw": "平均功率 (kW)",
    }

    current = data.get("current", {})
    previous = data.get("previous", {})
    mom = data.get("mom_change_pct", {})
    yoy = data.get("yoy_change_pct", {})

    row_idx = header_row + 1
    for key, label in metric_labels.items():
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=current.get(key))
        ws.cell(row=row_idx, column=3, value=previous.get(key))
        ws.cell(row=row_idx, column=4, value=mom.get(key))
        ws.cell(row=row_idx, column=5, value=yoy.get(key))
        row_idx += 1

    end_row = row_idx - 1
    _style_data_cells(ws, header_row + 1, end_row, len(COMPARISON_HEADERS))

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return _to_bytes(wb)

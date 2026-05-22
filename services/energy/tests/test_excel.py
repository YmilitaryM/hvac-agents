"""Tests for Excel export functionality."""
import io

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient

from energy_service.main import app
from energy_service.excel_generator import (
    generate_energy_report_excel,
    generate_mv_excel,
    generate_comparison_excel,
)


XLSX_MAGIC = b"PK\x03\x04"


# ── Unit tests for Excel generators ──────────────────────────────────────

class TestGenerateEnergyReportExcel:
    def test_produces_valid_xlsx_bytes(self):
        data = {
            "items": [
                {"id": 1, "period": "day", "report_type": "daily", "created_at": "2026-05-20T08:00:00"},
                {"id": 2, "period": "month", "report_type": "audit", "created_at": "2026-05-01T08:00:00"},
            ],
        }
        result = generate_energy_report_excel("month", data)
        assert result.startswith(XLSX_MAGIC)

    def test_can_be_opened_by_openpyxl(self):
        import openpyxl

        data = {
            "items": [
                {"id": 1, "period": "day", "report_type": "daily", "created_at": "2026-05-20T08:00:00"},
                {"id": 2, "period": "month", "report_type": "audit", "created_at": "2026-05-01T08:00:00"},
            ],
        }
        result = generate_energy_report_excel("month", data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws is not None

    def test_contains_chinese_headers(self):
        data = {
            "items": [
                {"id": 1, "period": "day", "report_type": "daily", "created_at": "2026-05-20T08:00:00"},
            ],
        }
        result = generate_energy_report_excel("month", data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        header_texts = "".join(str(cell.value or "") for cell in ws[1])
        # Should contain Chinese characters
        assert any("一" <= ch <= "鿿" for ch in header_texts)

    def test_numeric_values_preserved(self):
        data = {
            "items": [
                {"id": 42, "period": "day", "report_type": "daily", "created_at": "2026-05-20T08:00:00"},
            ],
        }
        result = generate_energy_report_excel("month", data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        all_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_values.extend(row)
        assert 42 in all_values


class TestGenerateMvExcel:
    def test_produces_valid_xlsx_bytes(self):
        data = {
            "plant_id": 1,
            "baseline_energy_kwh": 120000.0,
            "actual_energy_kwh": 108000.0,
            "savings_kwh": 12000.0,
            "savings_pct": 10.0,
            "uncertainty_pct": 8.5,
            "cv_rmse_pct": 15.2,
            "nmbe_pct": -1.8,
            "compliant_ashrae_g14": True,
            "compliant_gb28750": True,
            "coal_equivalent_tons": 4.8,
            "carbon_reduction_kg": 9600.0,
        }
        result = generate_mv_excel(data)
        assert result.startswith(XLSX_MAGIC)

    def test_contains_chinese_headers(self):
        data = {
            "plant_id": 1,
            "baseline_energy_kwh": 120000.0,
            "actual_energy_kwh": 108000.0,
            "savings_kwh": 12000.0,
            "savings_pct": 10.0,
            "uncertainty_pct": 8.5,
            "cv_rmse_pct": 15.2,
            "nmbe_pct": -1.8,
            "compliant_ashrae_g14": True,
            "compliant_gb28750": True,
            "coal_equivalent_tons": 4.8,
            "carbon_reduction_kg": 9600.0,
        }
        result = generate_mv_excel(data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        header_col = "".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1))
        assert any("一" <= ch <= "鿿" for ch in header_col)


class TestGenerateComparisonExcel:
    def test_produces_valid_xlsx_bytes(self):
        data = {
            "plant_id": 1,
            "period": "month",
            "current": {"total_kwh": 108000, "avg_cop": 5.2, "avg_power_kw": 450},
            "previous": {"total_kwh": 112000, "avg_cop": 5.0, "avg_power_kw": 467},
            "mom_change_pct": {"total_kwh": -3.6, "avg_cop": 4.0, "avg_power_kw": -3.6},
            "yoy_change_pct": {"total_kwh": -5.2, "avg_cop": 6.1, "avg_power_kw": -5.2},
        }
        result = generate_comparison_excel(data)
        assert result.startswith(XLSX_MAGIC)

    def test_contains_chinese_headers(self):
        data = {
            "plant_id": 1,
            "period": "month",
            "current": {"total_kwh": 108000, "avg_cop": 5.2, "avg_power_kw": 450},
            "previous": {"total_kwh": 112000, "avg_cop": 5.0, "avg_power_kw": 467},
            "mom_change_pct": {"total_kwh": -3.6, "avg_cop": 4.0, "avg_power_kw": -3.6},
            "yoy_change_pct": {"total_kwh": -5.2, "avg_cop": 6.1, "avg_power_kw": -5.2},
        }
        result = generate_comparison_excel(data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        header_texts = "".join(str(cell.value or "") for cell in ws[1])
        assert any("一" <= ch <= "鿿" for ch in header_texts)


# ── Integration tests for download endpoints ─────────────────────────────

@pytest_asyncio.fixture
async def client():
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


@pytest.mark.asyncio
async def test_reports_download_endpoint_content_type(client):
    r = await client.get("/api/energy/reports/download?plant_id=1")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_mv_download_endpoint_content_type(client):
    r = await client.get("/api/energy/mv/download?plant_id=1")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_comparison_download_endpoint_content_type(client):
    r = await client.get("/api/energy/comparison/download?plant_id=1&period=month")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_reports_download_returns_valid_xlsx(client):
    r = await client.get("/api/energy/reports/download?plant_id=1")
    assert r.status_code == 200
    assert r.content.startswith(XLSX_MAGIC)


@pytest.mark.asyncio
async def test_mv_download_returns_valid_xlsx(client):
    r = await client.get("/api/energy/mv/download?plant_id=1")
    assert r.status_code == 200
    assert r.content.startswith(XLSX_MAGIC)


@pytest.mark.asyncio
async def test_comparison_download_returns_valid_xlsx(client):
    r = await client.get("/api/energy/comparison/download?plant_id=1&period=month")
    assert r.status_code == 200
    assert r.content.startswith(XLSX_MAGIC)
